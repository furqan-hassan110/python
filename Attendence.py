import cv2
import tkinter as tk
from datetime import datetime, timedelta
import openpyxl

class QRScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Scanner App")

        self.qr_mapping_class9 = {
            "F9-01": {"name": "Muhammad Behjat"},
            "F9-02": {"name": "Hussain Ali"},
            "F9-03": {"name": "Ali Akbar"},
            "F9-04": {"name": "Aun Raza"},
            "F9-05": {"name": "Abbas Haider"},
            # Add more QR code mappings for Class 9 as needed
        }

        self.qr_mapping_matric = {
            "F8-01": {"name": "Aabis Ali"},
            "F8-02": {"name": "Shaheer Raza"},
            "F8-03": {"name": "S.M.Mehdi"},
            "F8-04": {"name": "Hammad Raza"},
            "F8-05": {"name": "Alishan Raza"},
            # Add more QR code mappings for Matric as needed
        }

        self.users_data = set()
        self.camera_active = False

        self.scan_button = tk.Button(self.root, text="Scan QR Code", command=self.toggle_camera)
        
        self.scan_button.pack(pady=90)
        self.scan_button.pack(padx=90)
        # self.text= tk.Text("The Five attendence")
        # self.text.pack(pady=100)

        self.cap = cv2.VideoCapture(0)

        self.workbook_class9 = None
        self.file_path_class9 = None
        self.workbook_matric = None
        self.file_path_matric = None

    def toggle_camera(self):
        if not self.camera_active:
            self.scan_button.config(text="Stop Scanning")
            self.camera_active = True
            self.scan_qr_code()
        else:
            self.scan_button.config(text="Scan QR Code")
            self.camera_active = False

    def scan_qr_code(self):
        while self.camera_active:
            ret, frame = self.cap.read()

            detector = cv2.QRCodeDetector()
            data, _, _ = detector.detectAndDecode(frame)

            if data:
                self.process_qr_data(data)

            cv2.imshow("QR Code Scanner", frame)

            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

            self.root.update()
            self.root.after(100)

        cv2.destroyAllWindows()

    def process_qr_data(self, data):
        if data in self.qr_mapping_class9:
            self.process_data(data, "Class 9")
        elif data in self.qr_mapping_matric:
            self.process_data(data, "Matric")

    def process_data(self, qr_data, student_class):
        if qr_data not in self.users_data:
            self.users_data.add(qr_data)

            workbook, file_path = self.get_or_create_workbook(student_class)

            if workbook.active['A1'].value is None:
                self.initialize_workbook_headers(workbook, student_class)

            today = datetime.today()
            day_column = self.get_day_column(today)
            student_name = self.get_student_name(qr_data, student_class)
            row_num = self.get_student_row_number(student_name, student_class)

            # Mark 'P' in bold and green color
            cell = workbook.active.cell(row=row_num, column=day_column, value='P')
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            font = openpyxl.styles.Font(color='00FF00', bold=True)
            cell.font = font

            workbook.save(file_path)
            print(f"Attendance marked for {student_name} on {today.strftime('%Y-%m-%d')} in {file_path}")

    def initialize_workbook_headers(self, workbook, student_class):
        workbook.active['A1'] = 'Student Name'

        # Set an appropriate width for the first column
        workbook.active.column_dimensions['A'].width = 20

        start_date = datetime(2023, 12, 7)
        for day in range((datetime(2024, 12, 7) - start_date).days + 1):
            day_column = day + 2
            header_date = start_date + timedelta(days=day)
            header_value = header_date.strftime("%Y-%m-%d") + f" ({self.get_day_name(header_date)})"
            workbook.active.cell(row=1, column=day_column, value=header_value)

        for row_num, name in enumerate(self.get_student_names(student_class), start=2):
            workbook.active.cell(row=row_num, column=1, value=name)

    def get_day_name(self, date):
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        return day_names[date.weekday()]

    def get_day_column(self, today):
        start_date = datetime(2023, 12, 7)
        return (today - start_date).days + 2

    def get_student_name(self, qr_data, student_class):
        if student_class == "Class 9":
            return self.qr_mapping_class9[qr_data]["name"]
        elif student_class == "Matric":
            return self.qr_mapping_matric[qr_data]["name"]

    def get_student_row_number(self, student_name, student_class):
        for row_num, name in enumerate(self.get_student_names(student_class), start=2):
            if name == student_name:
                return row_num

    def get_student_names(self, student_class):
        if student_class == "Class 9":
            return [data["name"] for data in self.qr_mapping_class9.values()]
        elif student_class == "Matric":
            return [data["name"] for data in self.qr_mapping_matric.values()]

    def get_or_create_workbook(self, class_name):
        if class_name == "Class 9":
            if self.workbook_class9 is None:
                self.workbook_class9, self.file_path_class9 = self.create_workbook(class_name)
            return self.workbook_class9, self.file_path_class9
        elif class_name == "Matric":
            if self.workbook_matric is None:
                self.workbook_matric, self.file_path_matric = self.create_workbook(class_name)
            return self.workbook_matric, self.file_path_matric

    def create_workbook(self, class_name):
        workbook = openpyxl.Workbook()
        return workbook, f"{class_name}_attendance.xlsx"


if __name__ == "__main__":
    root = tk.Tk()
    app = QRScannerApp(root)
    # root.protocol("WM_DELETE_WINDOW", app.toggle_camera)
    root.mainloop()
